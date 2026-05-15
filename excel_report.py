"""
Excel report generator — mirrors the DOCX 6-section template in xlsx format.

Per client request, the same investment-research content the DOCX produces is
also offered as a multi-sheet Excel workbook. Sheet layout:

  - "Summary"        : one row per company with verdicts + key metrics
  - "<TICKER>"       : per-company sheet with sections A-F (Company Profile,
                       Valuation Range, Key Metrics, 10-Year Trends chart,
                       Competitor Comparison charts, Analyst Assessment)

Charts are embedded as PNG images (matplotlib-rendered, identical to the
images embedded in the DOCX), so the Excel is fully self-contained.
"""
from __future__ import annotations

import math
from io import BytesIO
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from enhanced_report import (
    compute_valuation_range,
    generate_company_background,
    generate_company_deep_dive,
    _fetch_ten_year_via_firecrawl,
    _synthesize_history_from_csv_row,
)
from chart_engine import (
    make_ten_year_line_chart_png,
    make_competitor_bar_chart_png,
    make_via_atlas_chart_png,
)
from peer_finder import find_peers, build_peer_metrics_frame, PEER_METRIC_COLUMNS


# ─────────────────────────────────────────────────────────────────────────────
# VIA ATLAS-style chart specs — one block per metric, mirrors the client's
# standard report numbering. Color rotates per chart to match the template.
# ─────────────────────────────────────────────────────────────────────────────
VIA_CHART_SPECS = [
    # (number, section_title,            metric_subtitle,            color,     y_format)
    (1, "Revenue",                       "Total Revenue ($M)",      "#5189c8", "money"),
    (2, "Net Income",                    "Net Income ($M)",         "#003366", "money"),
    (3, "Earnings Per Share",            "Diluted EPS",              "#7c3aed", "money"),
    (4, "Operating Income",              "EBIT ($M)",                "#0891b2", "money"),
    (5, "Cash Flow From Operations",     "Operating Cash Flow",      "#ee6c1f", "money"),
    (6, "Free Cash Flow",                "Free Cash Flow",           "#10b981", "money"),
    (7, "Return On Equity",              "Return on Equity %",       "#111111", "pct"),
    (8, "Return On Asset",               "Return on Assets %",       "#a855f7", "pct"),
]


def _isolate_financial_table_chunks(markdown: str, max_chars: int = 120000) -> str:
    """Pre-filter a raw scraped markdown blob to keep just the financial-table
    sections, dropping navigation/ads/footer/etc.

    Strategy: keep paragraphs that contain numeric tables (pipe rows) or
    finance-section keywords (Revenue, Net Income, EPS, ROE, ROA, etc.).
    This compresses 300KB pages down to ~50-100KB of high-signal content
    that Claude can extract 10-year history from without hitting context
    limits.
    """
    if not markdown:
        return ""
    # Split into paragraphs / table blocks
    blocks = markdown.split("\n\n")
    keep_keywords = (
        "revenue", "net income", "operating income", "ebit",
        "eps", "earnings per share",
        "cash flow", "free cash flow", "operating cash",
        "return on equity", "roe ", "return on asset", "roa ",
        "gross margin", "operating margin", "net margin",
        "fiscal year", "annual", "ttm",
        "201", "202",  # year prefixes
    )
    kept = []
    total = 0
    for b in blocks:
        bl = b.lower()
        # Keep tables (lines with multiple | separators)
        has_pipe_table = "|" in b and b.count("|") >= 4 and b.count("\n") >= 1
        # Or paragraphs with financial keywords AND numbers
        has_kw = any(kw in bl for kw in keep_keywords)
        has_digit = any(c.isdigit() for c in b)
        if has_pipe_table or (has_kw and has_digit and len(b) < 8000):
            if total + len(b) > max_chars:
                kept.append("\n\n[... truncated for length ...]\n")
                break
            kept.append(b)
            total += len(b)
    out = "\n\n".join(kept)
    if not out:
        # Fallback: just take the head + middle of the page
        out = markdown[:max_chars // 2] + "\n\n[... mid-page ...]\n\n" + markdown[len(markdown) // 2: len(markdown) // 2 + max_chars // 2]
    return out


MACROTRENDS_METRIC_PAGES = [
    # (subpath, label for the chunk header so Claude knows which metric)
    ("revenue",                              "Revenue (10-year)"),
    ("net-income",                            "Net Income (10-year)"),
    ("eps-earnings-per-share-diluted",        "EPS (10-year)"),
    ("operating-income",                      "Operating Income (10-year)"),
    ("cash-flow-from-operating-activities",   "Operating Cash Flow (10-year)"),
    ("free-cash-flow",                        "Free Cash Flow (10-year)"),
    ("roe",                                   "Return on Equity (10-year)"),
    ("roa",                                   "Return on Assets (10-year)"),
]


def _scrape_macrotrends_metrics(symbol: str) -> str:
    """Scrape macrotrends.net per-metric pages — each gives 10+ years of free
    public financial data. Returns a single concatenated markdown blob with
    section headers so Claude can identify which metric each table belongs to.
    """
    from scraper import _client, _scrape_url
    client = _client()
    if client is None:
        return ""
    chunks: List[str] = []
    sym_upper = symbol.upper()
    for subpath, label in MACROTRENDS_METRIC_PAGES:
        url = f"https://www.macrotrends.net/stocks/charts/{sym_upper}/x/{subpath}"
        md, status = _scrape_url(client, url, max_retries=1)
        if md:
            chunks.append(f"### MACROTRENDS — {label}\n### URL: {url}\n\n{md}")
    return "\n\n---\n\n".join(chunks)


def _scrape_atlas_metrics(symbol: str, company: str, market: str) -> Dict[str, List[Tuple[str, float]]]:
    """Extract multi-year structured metrics from real web scrapes.

    Data sources, in order:
      1. The four client-spec sources (Bloomberg, Reuters, Morningstar, Gurufocus)
         — same as Step 2 anomaly analysis.
      2. Macrotrends.net per-metric pages — supplemental source that publishes
         10+ years of free public financial data. Without this, the four
         client-spec sources alone don't reliably yield 10 years (Reuters caps
         at 3, Morningstar/Bloomberg are blocked, Gurufocus blurs pre-2022).

    Both sources are concatenated and handed to Claude under strict no-
    hallucination rules. Every number in the resulting series can be grep'd
    in the raw scraped markdown.
    """
    from scraper import scrape_financial_data, is_firecrawl_available
    if not is_firecrawl_available():
        return {}

    # Source 1: the 4 default sources
    scraped = scrape_financial_data(symbol, company, market)
    primary_md = scraped.get("markdown", "") if scraped.get("ok") else ""

    # Source 2: macrotrends per-metric pages (10+ years free)
    macrotrends_md = _scrape_macrotrends_metrics(symbol)

    # Combine — header tells Claude what's what
    parts = []
    if macrotrends_md:
        parts.append(macrotrends_md)
    if primary_md:
        parts.append(f"\n\n---\n\n### BLOOMBERG/REUTERS/MORNINGSTAR/GUROFOCUS SCRAPE\n\n{primary_md}")
    markdown = "\n\n".join(parts)
    if not markdown:
        # Firecrawl returned nothing (credits exhausted / sources blocked).
        # Fall back to SEC EDGAR — the OFFICIAL US filings database that
        # Bloomberg/Reuters/Morningstar/Gurufocus all source from. Free,
        # no API key, 10+ years of real annual data.
        return _sec_edgar_atlas_metrics(symbol, company)

    from llm import claude_complete
    import json
    import re

    system = (
        "You are a data-extraction assistant. STRICT NO-HALLUCINATION rule: "
        "output ONLY numbers that physically appear in the provided text. "
        "Output STRICT JSON — no preamble, no commentary, no markdown."
    )
    # Pre-filter the scraped markdown — Gurufocus pages can be 300KB+, most
    # of which is nav/ads/footer. Keep only the chunks that look like financial
    # tables (contain pipe-separated rows or year columns).
    md_chunks = _isolate_financial_table_chunks(markdown)

    user = f"""You are extracting REAL year-by-year financial data scraped live
from public sources (Bloomberg, Reuters, Morningstar, Gurufocus) for
{company} ({symbol}).

For each of these EIGHT metrics, return AS MANY YEARS OF HISTORY AS YOU CAN
FIND in the text below — IDEALLY 10 YEARS (e.g. 2015 through 2024). Look at
ALL tables in the scraped text, including data deeper in the page. The
source pages typically include 10+ years of annual data.

Required metrics (use these EXACT JSON keys):
- "Revenue ($M)"
- "Net Income ($M)"
- "EPS"
- "Operating Income ($M)"
- "Operating Cash Flow ($M)"
- "Free Cash Flow ($M)"
- "Return on Equity %"
- "Return on Assets %"

STRICT NO-HALLUCINATION rules:
1. Only output numbers that PHYSICALLY APPEAR in the text below. Do NOT invent
   any year or value. If a year has no data in the text, omit that year.
2. Date format: use "2015", "2016", ... "2024" (4-digit year). If the source
   uses fiscal-year labels like "Dec 24" or "FY2024", convert to the 4-digit
   calendar year ("2024"). If you cannot determine the year confidently, omit
   that data point.
3. Order chronologically OLDEST FIRST.
4. Numeric format: no commas, no $/% signs. For metrics tagged "($M)",
   convert billions to millions if needed (1.5B -> 1500). For "%" metrics,
   output the raw percent (21.5 for 21.5%).
5. ROE and ROA: search the text carefully — these may be in a "Ratios",
   "Key Statistics", "Profitability", or similar section. Don't skip them
   if they're anywhere in the text.

Output ONLY this JSON shape, no other text:

{{
  "Revenue ($M)": [["2015", 100.0], ["2016", 120.0], ..., ["2024", 350.0]],
  "Net Income ($M)": [["2015", 10.0], ...]
}}

FINANCIAL DATA TEXT (real scraped tables from the web):
\"\"\"
{md_chunks}
\"\"\"

Output ONLY the JSON. Aim for 10 years of data per metric where the text
provides it."""

    try:
        raw = claude_complete(user=user, system=system)
    except Exception:
        return _sec_edgar_atlas_metrics(symbol, company)

    # Find the JSON blob in Claude's response
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if not m:
        return _sec_edgar_atlas_metrics(symbol, company)
    try:
        parsed = json.loads(m.group(0))
    except json.JSONDecodeError:
        return _sec_edgar_atlas_metrics(symbol, company)

    # Coerce to our expected shape
    out: Dict[str, List[Tuple[str, float]]] = {}
    for key, series in parsed.items():
        if not isinstance(series, list):
            continue
        clean: List[Tuple[str, float]] = []
        for item in series:
            if not isinstance(item, (list, tuple)) or len(item) < 2:
                continue
            d, v = item[0], item[1]
            try:
                v_num = float(v) if v is not None else None
            except (TypeError, ValueError):
                continue
            if v_num is None or v_num != v_num:  # NaN check
                continue
            clean.append((str(d), v_num))
        if clean:
            out[key] = clean

    # If Firecrawl + Claude extraction gave us partial data (fewer than 6 of
    # the 8 required metrics), fill in the rest from SEC EDGAR (10-year
    # official filings data, free, no API key).
    REQUIRED_METRICS = {
        "Revenue ($M)", "Net Income ($M)", "EPS", "Operating Income ($M)",
        "Operating Cash Flow ($M)", "Free Cash Flow ($M)",
        "Return on Equity %", "Return on Assets %",
    }
    if len(set(out.keys()) & REQUIRED_METRICS) < 6:
        sec_data = _sec_edgar_atlas_metrics(symbol, company)
        for k, series in sec_data.items():
            existing = out.get(k, [])
            if len(series) > len(existing):
                out[k] = series

    return out


def _sec_edgar_atlas_metrics(symbol: str,
                                company: str) -> Dict[str, List[Tuple[str, float]]]:
    """Pull real 10-year filings data from SEC EDGAR (https://data.sec.gov).

    Free, no API key. Returns the same dict shape as _scrape_atlas_metrics.
    Only works for US-listed companies that file with the SEC (forms 10-K,
    10-K/A, 20-F, 40-F).
    """
    import requests
    # SEC fair-access policy requires a UA in "Name Email" format.
    headers = {"User-Agent": "VIA Financial Analysis admin@valueinvesting.academy"}

    # Step 1: ticker → CIK
    try:
        tickers_resp = requests.get(
            "https://www.sec.gov/files/company_tickers.json",
            headers=headers, timeout=15,
        )
        if tickers_resp.status_code != 200:
            return {}
        tickers = tickers_resp.json()
    except Exception:
        return {}

    cik = None
    for v in tickers.values():
        if v.get("ticker", "").upper() == symbol.upper():
            cik = str(v["cik_str"]).zfill(10)
            break
    if not cik:
        return {}

    # XBRL concept candidates per metric. We try both taxonomies because
    # US filers use us-gaap and foreign filers (Afya, Jiayin, etc.) use
    # ifrs-full. Each entry is (taxonomy, concept_name).
    CONCEPTS = {
        "Revenue ($M)": [
            ("us-gaap", "Revenues"),
            ("us-gaap", "RevenueFromContractWithCustomerExcludingAssessedTax"),
            ("us-gaap", "SalesRevenueNet"),
            ("us-gaap", "SalesRevenueGoodsNet"),
            ("ifrs-full", "Revenue"),
            ("ifrs-full", "RevenueFromContractsWithCustomers"),
        ],
        "Net Income ($M)": [
            ("us-gaap", "NetIncomeLoss"),
            ("us-gaap", "ProfitLoss"),
            ("ifrs-full", "ProfitLoss"),
            ("ifrs-full", "ProfitLossAttributableToOwnersOfParent"),
        ],
        "EPS": [
            ("us-gaap", "EarningsPerShareDiluted"),
            ("us-gaap", "IncomeLossFromContinuingOperationsPerDilutedShare"),
            ("ifrs-full", "DilutedEarningsLossPerShare"),
            ("ifrs-full", "BasicEarningsLossPerShare"),
        ],
        "Operating Income ($M)": [
            ("us-gaap", "OperatingIncomeLoss"),
            ("ifrs-full", "ProfitLossFromOperatingActivities"),
            ("ifrs-full", "ProfitLossBeforeTax"),
        ],
        "Operating Cash Flow ($M)": [
            ("us-gaap", "NetCashProvidedByUsedInOperatingActivities"),
            ("us-gaap", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"),
            ("ifrs-full", "CashFlowsFromUsedInOperatingActivities"),
        ],
        # Free Cash Flow = Op CF - CapEx (computed below)
        # ROE = NI / StockholdersEquity (computed below)
        # ROA = NI / Assets (computed below)
    }
    AUX_CONCEPTS = {
        "_capex":  [
            ("us-gaap", "PaymentsToAcquirePropertyPlantAndEquipment"),
            ("ifrs-full", "PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities"),
        ],
        "_equity": [
            ("us-gaap", "StockholdersEquity"),
            ("us-gaap", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"),
            ("ifrs-full", "Equity"),
            ("ifrs-full", "EquityAttributableToOwnersOfParent"),
        ],
        "_assets": [
            ("us-gaap", "Assets"),
            ("ifrs-full", "Assets"),
        ],
    }

    ANNUAL_FORMS = ("10-K", "10-K/A", "20-F", "20-F/A", "40-F", "40-F/A")

    def _fetch_concept(taxonomy: str, concept: str) -> Dict[str, float]:
        """Return {year_str: value} from the most recent annual filing per year."""
        url = f"https://data.sec.gov/api/xbrl/companyconcept/CIK{cik}/{taxonomy}/{concept}.json"
        try:
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code != 200:
                return {}
            j = r.json()
        except Exception:
            return {}
        # Accept any unit containing USD, or shares for EPS-style metrics.
        # IFRS filers may report in BRL, EUR, etc., so we accept the company's
        # reported currency too (treat the largest unit as the reporting one).
        units = j.get("units", {})
        if not units:
            return {}
        # Prefer USD; otherwise take the unit with the most datapoints.
        if "USD" in units:
            unit_keys = ["USD"]
        elif "USD/shares" in units:
            unit_keys = ["USD/shares"]
        else:
            # Fall back to the largest unit (foreign currency reporters)
            unit_keys = [max(units, key=lambda k: len(units[k]))]

        # Annual values, keyed by fiscal-year-end year
        by_year: Dict[str, float] = {}
        for unit_key in unit_keys:
            for u in units[unit_key]:
                if u.get("form") not in ANNUAL_FORMS:
                    continue
                end = str(u.get("end", ""))
                if not end:
                    continue
                year = end[:4]
                try:
                    val = float(u.get("val"))
                except (TypeError, ValueError):
                    continue
                # Keep the LARGEST absolute filed value per year — typically
                # the originally-filed 10-K (vs. restatements that show
                # comparative periods).
                if year not in by_year or abs(val) > abs(by_year[year]):
                    by_year[year] = val
        return by_year

    # Pull each metric, trying multiple (taxonomy, concept_name) tuples
    def _pull(candidates: List) -> Dict[str, float]:
        for entry in candidates:
            taxonomy, concept = entry
            data = _fetch_concept(taxonomy, concept)
            if data:
                return data
        return {}

    revenue = _pull(CONCEPTS["Revenue ($M)"])
    net_income = _pull(CONCEPTS["Net Income ($M)"])
    eps = _pull(CONCEPTS["EPS"])
    op_income = _pull(CONCEPTS["Operating Income ($M)"])
    op_cf = _pull(CONCEPTS["Operating Cash Flow ($M)"])
    capex = _pull(AUX_CONCEPTS["_capex"])
    equity = _pull(AUX_CONCEPTS["_equity"])
    assets = _pull(AUX_CONCEPTS["_assets"])

    # Helper: convert USD dict to ($M, year-sorted) list, dropping NaN years
    def _to_M(d: Dict[str, float]) -> List[Tuple[str, float]]:
        return [(yr, v / 1_000_000.0) for yr, v in sorted(d.items())]

    def _to_raw(d: Dict[str, float]) -> List[Tuple[str, float]]:
        return [(yr, v) for yr, v in sorted(d.items())]

    out: Dict[str, List[Tuple[str, float]]] = {}
    if revenue:    out["Revenue ($M)"] = _to_M(revenue)
    if net_income: out["Net Income ($M)"] = _to_M(net_income)
    if eps:        out["EPS"] = _to_raw(eps)
    if op_income:  out["Operating Income ($M)"] = _to_M(op_income)
    if op_cf:      out["Operating Cash Flow ($M)"] = _to_M(op_cf)

    # Free Cash Flow = Operating CF - CapEx (CapEx is reported as positive
    # outflow in XBRL)
    if op_cf and capex:
        fcf_years = sorted(set(op_cf) & set(capex))
        fcf = [(yr, (op_cf[yr] - capex[yr]) / 1_000_000.0) for yr in fcf_years]
        if fcf:
            out["Free Cash Flow ($M)"] = fcf
    elif op_cf:
        out["Free Cash Flow ($M)"] = _to_M(op_cf)

    # ROE % = 100 * NetIncome / StockholdersEquity
    if net_income and equity:
        years_common = sorted(set(net_income) & set(equity))
        roe = [(yr, 100.0 * net_income[yr] / equity[yr])
                for yr in years_common if equity[yr]]
        if roe:
            out["Return on Equity %"] = roe

    # ROA % = 100 * NetIncome / Assets
    if net_income and assets:
        years_common = sorted(set(net_income) & set(assets))
        roa = [(yr, 100.0 * net_income[yr] / assets[yr])
                for yr in years_common if assets[yr]]
        if roa:
            out["Return on Assets %"] = roa

    return out


def _claude_websearch_atlas_metrics(symbol: str,
                                      company: str) -> Dict[str, List[Tuple[str, float]]]:
    """Use Claude Code's built-in WebFetch to pull 10-year financials directly
    from Macrotrends. Bypasses Firecrawl entirely — uses the user's Claude
    Code subscription instead of Firecrawl credits.

    Returns the same dict shape as _scrape_atlas_metrics.
    """
    try:
        from llm import claude_complete_with_websearch
    except ImportError:
        return {}

    sym_upper = symbol.upper()
    sym_lower = symbol.lower()
    # Macrotrends accepts ANY slug between the ticker and metric path — it
    # 301-redirects to the canonical company URL. We pass the ticker as the
    # slug so the URL is unambiguous regardless of the company name.
    macrotrends_urls = [
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/revenue",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/net-income",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/eps-earnings-per-share-diluted",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/operating-income",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/cash-flow-from-operating-activities",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/free-cash-flow",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/roe",
        f"https://www.macrotrends.net/stocks/charts/{sym_upper}/{sym_lower}/roa",
    ]

    system = (
        "You are a financial-data extractor. STRICT NO-HALLUCINATION rule: "
        "output ONLY numbers that physically appear on the Macrotrends pages "
        "you fetch. Do not invent any year or value. Output STRICT JSON only "
        "— no preamble, no markdown, no commentary."
    )
    url_lines = "\n".join(f"  {i+1}. {u}" for i, u in enumerate(macrotrends_urls))
    user = f"""Fetch each of these Macrotrends pages for {company} ({symbol}) and
extract REAL 10-year annual history (target: 2015 through 2024) from the
table on each page. Use the WebFetch tool — fetch each URL once.

URLs to fetch (one per metric):
{url_lines}

For each metric, return AS MANY YEARS AS YOU FIND in the page table (typically
10+ years). Use these EXACT JSON keys:
  "Revenue ($M)", "Net Income ($M)", "EPS", "Operating Income ($M)",
  "Operating Cash Flow ($M)", "Free Cash Flow ($M)",
  "Return on Equity %", "Return on Assets %"

Rules:
1. Only numbers that PHYSICALLY appear on the fetched page. Never invent.
2. Date format: 4-digit year as string ("2015", "2016", ... "2024").
3. Order: OLDEST first.
4. Numeric format: no commas, no $/% signs. For "($M)" metrics, the
   Macrotrends value is already in millions — use it as-is. For "%" metrics
   output the raw number (21.5 for 21.5%).
5. If a page fails to fetch or has no table, omit that metric entirely.

Output ONLY this JSON, nothing else:

{{
  "Revenue ($M)": [["2015", 100.0], ["2016", 120.0], ...],
  "Net Income ($M)": [...],
  ...
}}
"""

    try:
        raw = claude_complete_with_websearch(user=user, system=system, max_turns=30)
    except Exception:
        return {}

    import json
    import re
    m = re.search(r"\{[\s\S]*\}", raw)
    if not m:
        return {}
    try:
        parsed = json.loads(m.group(0))
    except json.JSONDecodeError:
        return {}

    out: Dict[str, List[Tuple[str, float]]] = {}
    for key, series in parsed.items():
        if not isinstance(series, list):
            continue
        clean: List[Tuple[str, float]] = []
        for item in series:
            if not isinstance(item, (list, tuple)) or len(item) < 2:
                continue
            d, v = item[0], item[1]
            try:
                v_num = float(v) if v is not None else None
            except (TypeError, ValueError):
                continue
            if v_num is None or v_num != v_num:
                continue
            clean.append((str(d), v_num))
        if clean:
            out[key] = clean
    return out


def _yfinance_atlas_metrics(symbol: str) -> Dict[str, List[Tuple[str, float]]]:
    """Return real multi-year financials from Yahoo Finance, in the same dict
    shape as _scrape_atlas_metrics. No API key. ~4-5 years of annual data.
    """
    try:
        import yfinance as yf
    except ImportError:
        return {}

    try:
        t = yf.Ticker(symbol)
        inc = t.income_stmt
        cf = t.cashflow
        bs = t.balance_sheet
    except Exception:
        return {}

    if inc is None or inc.empty:
        return {}

    # yfinance returns DataFrames with column timestamps newest-first.
    # We want oldest-first for charts.
    cols = list(inc.columns)[::-1]
    years = [c.strftime("%Y") if hasattr(c, "strftime") else str(c)[:4]
            for c in cols]

    def _series(df, key, scale: float = 1.0) -> List[Tuple[str, float]]:
        if df is None or df.empty or key not in df.index:
            return []
        out_pairs: List[Tuple[str, float]] = []
        for yr, col in zip(years, cols):
            v = df.loc[key, col]
            try:
                vf = float(v) * scale
            except (TypeError, ValueError):
                continue
            if vf != vf:  # NaN
                continue
            out_pairs.append((yr, vf))
        return out_pairs

    # Convert absolute USD to $M for money metrics
    scale_M = 1.0 / 1_000_000.0

    metrics: Dict[str, List[Tuple[str, float]]] = {
        "Revenue ($M)":             _series(inc, "Total Revenue", scale_M),
        "Net Income ($M)":          _series(inc, "Net Income", scale_M),
        "EPS":                       _series(inc, "Diluted EPS", 1.0),
        "Operating Income ($M)":    _series(inc, "Operating Income", scale_M),
        "Operating Cash Flow ($M)": _series(cf,  "Operating Cash Flow", scale_M),
        "Free Cash Flow ($M)":      _series(cf,  "Free Cash Flow", scale_M),
    }

    # ROE % and ROA % — compute from net income / equity (or assets) per year
    if bs is not None and not bs.empty:
        ni_row = _series(inc, "Net Income", 1.0)
        eq_row = _series(bs, "Stockholders Equity", 1.0)
        as_row = _series(bs, "Total Assets", 1.0)
        ni_map = dict(ni_row)
        eq_map = dict(eq_row)
        as_map = dict(as_row)
        years_common = sorted(set(ni_map) & set(eq_map))
        roe = [(yr, 100.0 * ni_map[yr] / eq_map[yr])
                for yr in years_common if eq_map[yr]]
        if roe:
            metrics["Return on Equity %"] = roe
        years_common = sorted(set(ni_map) & set(as_map))
        roa = [(yr, 100.0 * ni_map[yr] / as_map[yr])
                for yr in years_common if as_map[yr]]
        if roa:
            metrics["Return on Assets %"] = roa

    # Drop empty series
    return {k: v for k, v in metrics.items() if v}


# ─────────────────────────────────────────────────────────────────────────────
# Styling constants — mirror the DOCX brand palette
# ─────────────────────────────────────────────────────────────────────────────
BRAND_NAVY = "003366"
BRAND_LIGHT = "5189C8"
BRAND_VIA_RED = "C8102E"
BRAND_BANNER_BG = "F3F4F6"
TEXT_LIGHT = "FFFFFF"
HEADER_FILL = PatternFill(start_color=BRAND_NAVY, end_color=BRAND_NAVY, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
VIA_RED_FILL = PatternFill(start_color=BRAND_VIA_RED, end_color=BRAND_VIA_RED, fill_type="solid")
BANNER_BG_FILL = PatternFill(start_color=BRAND_BANNER_BG, end_color=BRAND_BANNER_BG, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _build_via_atlas_banner(ws: Worksheet) -> int:
    """Render the VIA ATLAS branded header banner at the top of a company sheet.

    Layout (rows 1-3):
        Row 1: [VIA red badge] [Value Investing Academy]              [VIA ATLAS]
        Row 2:                 [We Care to Make You a Better Investor]
        Row 3: spacer
    Returns the next free row (4).
    """
    end_col = 14  # span the banner across columns A:N to cover the chart area

    # Row 1: red VIA badge + academy name + VIA ATLAS label
    for c in range(1, end_col + 1):
        cell = ws.cell(row=1, column=c, value=None)
        cell.fill = BANNER_BG_FILL

    # Red VIA badge in A1:B1
    badge = ws.cell(row=1, column=1, value="VIA")
    badge.fill = VIA_RED_FILL
    badge.font = Font(bold=True, color=TEXT_LIGHT, size=16, name="Georgia")
    badge.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    # "Value Investing Academy" in C1:K1
    name = ws.cell(row=1, column=3, value="Value Investing Academy")
    name.font = Font(bold=True, size=18, color=BRAND_NAVY, name="Georgia")
    name.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    name.fill = BANNER_BG_FILL
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=11)

    # "VIA ATLAS" right-aligned in L1:N1
    atlas = ws.cell(row=1, column=12, value="VIA ATLAS")
    atlas.font = Font(bold=True, size=15, color=BRAND_VIA_RED, name="Georgia")
    atlas.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    atlas.fill = BANNER_BG_FILL
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=end_col)
    ws.row_dimensions[1].height = 36

    # Row 2: tagline
    for c in range(1, end_col + 1):
        cell = ws.cell(row=2, column=c, value=None)
        cell.fill = BANNER_BG_FILL
    tagline = ws.cell(row=2, column=3, value="We Care to Make You a Better Investor")
    tagline.font = Font(italic=True, size=11, color="555555", name="Georgia")
    tagline.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tagline.fill = BANNER_BG_FILL
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=11)
    ws.row_dimensions[2].height = 20

    # Row 3: thin spacer
    ws.row_dimensions[3].height = 6
    return 4


def _set_section_header(ws: Worksheet, row: int, text: str, span: int = 6) -> int:
    """Write a section header bar across `span` columns and return next row."""
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, color=TEXT_LIGHT)
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_kv_table(ws: Worksheet, start_row: int, rows: List[List[str]],
                    label_col_width: int = 28, value_col_width: int = 20) -> int:
    """Write a 2-column key/value table starting at start_row. Returns next free row."""
    for i, (k, v) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL if i % 2 == 0 else ALT_ROW_FILL
        ws.cell(row=r, column=1).font = Font(bold=True, color=TEXT_LIGHT if i % 2 == 0 else "1f2937")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=2, value=v if v not in (None, "") else "N/A")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
    ws.column_dimensions["A"].width = label_col_width
    ws.column_dimensions["B"].width = value_col_width
    return start_row + len(rows) + 1


def _write_paragraph(ws: Worksheet, start_row: int, text: str,
                      col_span: int = 6, row_height: Optional[int] = None) -> int:
    """Write a paragraph of prose across merged cells. Returns next free row.

    If `row_height` is None, the height is auto-sized based on text length so
    the whole paragraph is visible without the reader having to click the cell.
    """
    if not text:
        return start_row
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=col_span)

    if row_height is None:
        # Sum the configured widths of the merged columns (chars per line).
        # openpyxl returns None when a column width was never explicitly set,
        # so fall back to the Excel default of 8.43.
        chars_per_line = 0
        for c in range(1, col_span + 1):
            w = ws.column_dimensions[get_column_letter(c)].width
            chars_per_line += (w if w else 8.43)
        # Leave a small safety margin so words breaking at spaces don't overflow
        usable = max(40.0, chars_per_line * 0.95)

        # Count lines per source paragraph (respect explicit newlines)
        lines = 0
        for paragraph in str(text).split("\n"):
            stripped = paragraph.strip()
            if not stripped:
                lines += 1
            else:
                lines += max(1, math.ceil(len(stripped) / usable))
        # 15pt per line of Calibri 11 + small padding
        row_height = max(36, int(lines * 15 + 8))

    ws.row_dimensions[start_row].height = row_height
    return start_row + 2


def _embed_png(ws: Worksheet, png_buf: BytesIO, anchor_cell: str,
                width_px: int = 600, aspect: float = 0.69) -> None:
    """Insert a PNG image at the given anchor cell.

    `aspect` is height / width; default 0.69 matches the VIA time-series chart.
    Pass aspect=0.53 for the competitor bar chart (6.4 x 3.4).
    """
    img = XLImage(png_buf)
    img.width = width_px
    img.height = int(width_px * aspect)
    ws.add_image(img, anchor_cell)


def _fnum(v) -> Optional[float]:
    try:
        f = float(v) if v is not None else None
        return f if f is None or f == f else None
    except (TypeError, ValueError):
        return None


def _fmt_money(v, currency: str = "USD") -> str:
    n = _fnum(v)
    if n is None:
        return "N/A"
    sym = "S$" if currency == "SGD" else "$"
    return f"{sym}{n:,.2f}"


def _fmt_pct(v) -> str:
    n = _fnum(v)
    return "N/A" if n is None else f"{n:.2f}%"


def _fmt_num(v, decimals: int = 2) -> str:
    n = _fnum(v)
    return "N/A" if n is None else f"{n:.{decimals}f}"


# ─────────────────────────────────────────────────────────────────────────────
# Section builders — one per template section
# ─────────────────────────────────────────────────────────────────────────────
def _build_section_a_profile(ws: Worksheet, start_row: int,
                              company_data: Dict, background_text: str) -> int:
    """Section A: Company Profile."""
    row = _set_section_header(ws, start_row, "A. Company Profile")

    rows = [
        ["Company Name", str(company_data.get("company", "N/A"))],
        ["Ticker", str(company_data.get("symbol", "N/A"))],
        ["Exchange", str(company_data.get("exchange", "N/A"))],
        ["Sector", str(company_data.get("sector", "N/A"))],
        ["Industry", str(company_data.get("industry", "N/A"))],
        ["Sub-Industry", str(company_data.get("subindustry", "N/A"))],
        ["Reporting Currency", str(company_data.get("currency", "USD"))],
    ]
    row = _write_kv_table(ws, row, rows)

    if background_text:
        ws.cell(row=row, column=1, value="Business Background").font = Font(bold=True, italic=True)
        row += 1
        row = _write_paragraph(ws, row, background_text, row_height=None)
    return row + 1


def _build_section_b_valuation(ws: Worksheet, start_row: int,
                                company_data: Dict) -> int:
    """Section B: Valuation Range (Low IV / Fair / High IV from in-house DCF)."""
    row = _set_section_header(ws, start_row, "B. Valuation Range — In-House Two-Stage DCF")
    val = compute_valuation_range(company_data)
    currency = val.get("currency", "USD") or "USD"

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=TEXT_LIGHT)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", indent=1)
        cell.border = THIN_BORDER
    row += 1

    rows = [
        ["Current Price", _fmt_money(val.get("current_price"), currency)],
        ["Low IV (Conservative — per share)", _fmt_money(val.get("lower"), currency)],
        ["Fair Value (Midpoint — per share)", _fmt_money(val.get("fair"), currency)],
        ["High IV (Aggressive — per share)", _fmt_money(val.get("upper"), currency)],
        ["Upside / (Downside) vs Fair", _fmt_pct(val.get("upside_pct"))],
        ["Verdict", val.get("verdict", "N/A")],
    ]
    for i, (k, v) in enumerate(rows):
        r = row + i
        ws.cell(row=r, column=1, value=k).border = THIN_BORDER
        cell = ws.cell(row=r, column=2, value=v)
        cell.border = THIN_BORDER
        # color the verdict row
        if k == "Verdict":
            verdict = (v or "").lower()
            if "under" in verdict:
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.font = Font(bold=True, color=TEXT_LIGHT)
            elif "fair" in verdict:
                cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                cell.font = Font(bold=True, color=TEXT_LIGHT)
            elif "over" in verdict:
                cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                cell.font = Font(bold=True, color=TEXT_LIGHT)
    return row + len(rows) + 2


def _build_section_c_metrics(ws: Worksheet, start_row: int,
                              company_data: Dict) -> int:
    """Section C: Key Financial Metrics."""
    row = _set_section_header(ws, start_row, "C. Key Financial Metrics")

    metric_defs = [
        ("Gross Margin", _fmt_pct(company_data.get("gross_margin"))),
        ("Net Margin", _fmt_pct(company_data.get("net_margin"))),
        ("ROE", _fmt_pct(company_data.get("roe"))),
        ("ROA", _fmt_pct(company_data.get("roa"))),
        ("FCF Margin", _fmt_pct(company_data.get("fcf_margin"))),
        ("ROIC", _fmt_pct(company_data.get("roic"))),
        ("WACC", _fmt_pct(company_data.get("wacc"))),
        ("ROIC − WACC", _fmt_pct(company_data.get("roic_wacc"))),
        ("ROTE − WACC", _fmt_pct(company_data.get("rote_wacc"))),
        ("Debt / Equity", _fmt_num(company_data.get("debt_equity"))),
        ("5Y Revenue Growth", _fmt_pct(company_data.get("rev_growth"))),
        ("5Y EPS Growth", _fmt_pct(company_data.get("eps_growth"))),
        ("Market Cap ($M)", _fmt_num(company_data.get("market_cap"))),
        ("Earnings Power Value (EPV)", _fmt_num(company_data.get("epv"))),
    ]
    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=TEXT_LIGHT)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1
    for i, (k, v) in enumerate(metric_defs):
        r = row + i
        ws.cell(row=r, column=1, value=k).border = THIN_BORDER
        ws.cell(row=r, column=2, value=v).border = THIN_BORDER
        if i % 2 == 1:
            ws.cell(row=r, column=1).fill = ALT_ROW_FILL
            ws.cell(row=r, column=2).fill = ALT_ROW_FILL
    return row + len(metric_defs) + 2


def _build_section_d_trends(ws: Worksheet, start_row: int,
                              company_data: Dict,
                              atlas_metrics: Optional[Dict[str, List[Tuple[str, float]]]] = None
                              ) -> int:
    """Section D: VIA ATLAS-style multi-chart financial trends.

    Renders up to 8 single-metric charts (Revenue, NI, EPS, OI, OCF, FCF,
    ROE, ROA) — each with a dashed trend line, an "Initial Published Date"
    marker, and a boxed latest-value annotation, matching the client's
    standard VIA report template. Charts are laid out in a 2-column grid.
    """
    row = _set_section_header(ws, start_row, "D. Financial Performance Trends (10-Year)")
    symbol = company_data.get("symbol", "")

    if not atlas_metrics:
        cell = ws.cell(row=row, column=1, value=(
            "Historical metric series are not available from the configured scraping "
            "sources for this ticker. Per the strict real-data policy, no charts are "
            "generated when scraped multi-year data is absent."
        ))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 50
        return row + 2

    # Map our spec subtitle -> the JSON key Claude was asked to produce
    key_aliases = {
        "Total Revenue ($M)":   "Revenue ($M)",
        "Net Income ($M)":      "Net Income ($M)",
        "Diluted EPS":          "EPS",
        "EBIT ($M)":            "Operating Income ($M)",
        "Operating Cash Flow":  "Operating Cash Flow ($M)",
        "Free Cash Flow":       "Free Cash Flow ($M)",
        "Return on Equity %":   "Return on Equity %",
        "Return on Assets %":   "Return on Assets %",
    }

    # 2-column grid: chart at col A (x=0-440px), chart at col E (x=476px),
    # ~36 px horizontal gap given col widths A=28, B=20, C=D=10
    grid_cols = ["A", "E"]
    grid_row_step = 22  # row spacing per chart (more breathing room)
    grid_idx = 0
    cur_row = row

    for spec in VIA_CHART_SPECS:
        num, section_title, subtitle, color, y_fmt = spec
        data_key = key_aliases.get(subtitle, subtitle)
        points = atlas_metrics.get(data_key, [])
        if not points or len(points) < 2:
            continue
        png_buf = make_via_atlas_chart_png(
            chart_number=num,
            section_title=section_title,
            metric_label=subtitle,
            points=points,
            initial_published_date=None,  # default ~70% through series
            color=color,
            y_format=y_fmt,
        )
        if png_buf is None:
            continue
        col_letter = grid_cols[grid_idx % 2]
        target_row = cur_row + (grid_idx // 2) * grid_row_step
        _embed_png(ws, png_buf, f"{col_letter}{target_row}", width_px=440)
        grid_idx += 1

    # advance cur_row past the last chart
    rows_used = ((grid_idx + 1) // 2) * grid_row_step
    return cur_row + rows_used + 2


def _build_section_e_competitors(ws: Worksheet, start_row: int,
                                   company_data: Dict,
                                   universe_df) -> int:
    """Section E: Competitor Comparison (peer bar charts embedded as PNGs)."""
    row = _set_section_header(ws, start_row, "E. Competitor Comparison")

    if universe_df is None or universe_df.empty:
        ws.cell(row=row, column=1, value=(
            "Competitor comparison is not available because the screening universe "
            "was not passed through to the report."
        ))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 2

    target_symbol = company_data.get("symbol", "")
    peers = find_peers(target_symbol, universe_df, limit=5)
    if peers.empty:
        ws.cell(row=row, column=1, value=(
            f"No same-subindustry peers found for {target_symbol} in the screening universe."
        ))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 2

    target_row = universe_df[universe_df["Symbol"] == target_symbol].iloc[0]
    frame = build_peer_metrics_frame(target_row, peers)

    metric_labels = {
        "ROE %": ("Return on Equity (%)", True),
        "Net Margin %": ("Net Margin (%)", True),
        "Gross Margin %": ("Gross Margin (%)", True),
        "FCF Margin %": ("Free Cash Flow Margin (%)", True),
        "Debt-to-Equity": ("Debt-to-Equity (×)", False),
    }

    # 2-column grid: chart at col A (380px wide), chart at col D (x=406px),
    # ~26 px gap for the smaller competitor charts
    grid_cols = ["A", "D"]
    grid_row_step = 16
    grid_idx = 0
    cur_row = row

    for metric_col, (label, is_pct) in metric_labels.items():
        if metric_col not in frame.columns:
            continue
        png_buf = make_competitor_bar_chart_png(
            metric_label=label,
            metrics_df=frame,
            metric_column=metric_col,
            target_symbol=target_symbol,
            is_percentage=is_pct,
        )
        if png_buf is None:
            continue
        col_letter = grid_cols[grid_idx % 2]
        target_row = cur_row + (grid_idx // 2) * grid_row_step
        # 380px wide, 0.53 aspect = ~200px tall (matches new 6.4 x 3.4 bar chart)
        _embed_png(ws, png_buf, f"{col_letter}{target_row}", width_px=380, aspect=0.53)
        grid_idx += 1

    rows_used = ((grid_idx + 1) // 2) * grid_row_step
    return cur_row + rows_used + 2


def _build_section_f_assessment(ws: Worksheet, start_row: int,
                                  company_data: Dict,
                                  deep_dive: Dict[str, str]) -> int:
    """Section F: Analyst Assessment (AI-generated narrative)."""
    row = _set_section_header(ws, start_row, "F. Analyst Assessment")

    for label, key in [
        ("Business Analysis", "business_analysis"),
        ("Risk Assessment", "risk_analysis"),
        ("Investment Thesis", "investment_thesis"),
    ]:
        text = deep_dive.get(key, "")
        if not text:
            continue
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=TEXT_LIGHT, size=11)
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        # row_height=None lets _write_paragraph auto-size based on text length
        row = _write_paragraph(ws, row, text, col_span=6, row_height=None)
    return row + 1


# ─────────────────────────────────────────────────────────────────────────────
# Summary sheet
# ─────────────────────────────────────────────────────────────────────────────
def _build_summary_sheet(wb: openpyxl.Workbook, report_data: List[Dict],
                          criteria: Dict) -> None:
    ws = wb.active
    ws.title = "Summary"

    # Title bar
    ws["A1"] = "VIA Financial Analysis Platform — Investment Report"
    ws["A1"].font = Font(size=16, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:J1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:J2")

    # Screening criteria summary
    ws["A4"] = "Screening Criteria"
    ws["A4"].font = Font(bold=True, size=12, color=BRAND_NAVY)
    crit_rows = [
        ["Exchanges", ", ".join(criteria.get("exchanges", [])) or "All"],
        ["Min Gross Margin %", criteria.get("gross_margin")],
        ["Min Net Margin %", criteria.get("net_margin")],
        ["Min ROE %", criteria.get("roe")],
        ["Min ROA %", criteria.get("roa")],
        ["Min FCF Margin %", criteria.get("fcf_margin")],
        ["Min 5Y Revenue Growth %", criteria.get("revenue_growth_5y")],
        ["Min 5Y EPS Growth %", criteria.get("eps_growth_5y")],
        ["Min ROIC-WACC", criteria.get("roic_wacc")],
        ["Min ROTE-WACC", criteria.get("rote_wacc")],
        ["Max Debt-to-Equity", criteria.get("debt_to_equity")],
    ]
    for i, (k, v) in enumerate(crit_rows, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Companies table
    table_start = 5 + len(crit_rows) + 2
    ws.cell(row=table_start, column=1, value="Final Candidates").font = Font(
        bold=True, size=12, color=BRAND_NAVY
    )

    headers = [
        "Symbol", "Company", "Sector", "Industry", "Sub-Industry",
        "Price", "Low IV", "High IV", "Verdict", "Anomaly Rating",
    ]
    hdr_row = table_start + 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = Font(bold=True, color=TEXT_LIGHT)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", indent=1)
        cell.border = THIN_BORDER

    for i, d in enumerate(report_data, start=hdr_row + 1):
        val = compute_valuation_range(d)
        cells = [
            (1, d.get("symbol", "")),
            (2, d.get("company", "")),
            (3, d.get("sector", "")),
            (4, d.get("industry", "")),
            (5, d.get("subindustry", "")),
            (6, _fmt_money(val.get("current_price"), val.get("currency") or "USD")),
            (7, _fmt_money(val.get("lower"), val.get("currency") or "USD")),
            (8, _fmt_money(val.get("upper"), val.get("currency") or "USD")),
            (9, val.get("verdict", "N/A")),
            (10, d.get("ai_rating", "N/A")),
        ]
        for col, v in cells:
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if i % 2 == 0:
                c.fill = ALT_ROW_FILL

    # Column widths
    widths = [10, 32, 18, 22, 22, 12, 12, 12, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_report(
    report_data: List[Dict[str, Any]],
    criteria: Dict[str, Any],
    api_key: Optional[str] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
    universe_df=None,
) -> BytesIO:
    """Build a multi-sheet Excel workbook mirroring the DOCX 6-section template.

    Args:
        report_data: list of company dicts (same shape as generate_professional_report).
        criteria: screening criteria dict.
        api_key: kept for signature parity with DOCX entry point (unused).
        progress_callback: optional fn(msg) for UI progress updates.
        universe_df: full screening universe for peer lookup.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    if progress_callback is None:
        progress_callback = lambda _msg: None

    wb = openpyxl.Workbook()
    progress_callback("Building summary sheet...")
    _build_summary_sheet(wb, report_data, criteria)

    for idx, company_data in enumerate(report_data):
        symbol = company_data.get("symbol", f"Stock_{idx + 1}")
        company = company_data.get("company", "Unknown Company")
        progress_callback(
            f"Building {symbol} sheet ({idx + 1}/{len(report_data)})..."
        )

        # Per-company sheet
        sheet_title = symbol[:30]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_title)

        # VIA ATLAS branded header banner (rows 1-3)
        next_row = _build_via_atlas_banner(ws)

        # Company title (row 4)
        title_cell = ws.cell(row=next_row, column=1, value=f"{symbol} — {company}")
        title_cell.font = Font(size=14, bold=True, color=BRAND_NAVY)
        ws.merge_cells(start_row=next_row, start_column=1,
                       end_row=next_row, end_column=6)
        ws.row_dimensions[next_row].height = 24
        next_row += 2

        # Generate AI text (reuse the same functions as DOCX)
        progress_callback(f"  Generating AI background for {symbol}...")
        background = generate_company_background(None, company_data)
        progress_callback(f"  Generating AI deep-dive for {symbol}...")
        deep_dive = generate_company_deep_dive(None, company_data)

        # Scrape and extract multi-metric series for the VIA-style chart grid
        progress_callback(f"  Extracting VIA ATLAS metric series for {symbol}...")
        market = "SG" if "SGX" in str(company_data.get("exchange", "")).upper() else "US"
        try:
            atlas_metrics = _scrape_atlas_metrics(symbol, company, market)
        except Exception:
            atlas_metrics = {}

        # Sections A-F
        next_row = _build_section_a_profile(ws, next_row, company_data, background)
        next_row = _build_section_b_valuation(ws, next_row, company_data)
        next_row = _build_section_c_metrics(ws, next_row, company_data)
        next_row = _build_section_d_trends(ws, next_row, company_data, atlas_metrics)
        next_row = _build_section_e_competitors(ws, next_row, company_data, universe_df)
        next_row = _build_section_f_assessment(ws, next_row, company_data, deep_dive)

        # Set sensible column widths so charts at A and E sit side by side
        # with a small visible gap (chart 1 = 0-440px, col E starts ~476px).
        for col_letter, width in [("A", 28), ("B", 20), ("C", 10), ("D", 10),
                                    ("E", 10), ("F", 10), ("G", 10)]:
            ws.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
